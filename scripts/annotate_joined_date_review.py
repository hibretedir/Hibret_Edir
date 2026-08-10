#!/usr/bin/env python3
"""
Match joining dates from Excel → CRM members.review.xlsx for board review.
Does NOT update the database.

Default input: data/Members List as of 11-01-2024 Including Original.xlsx
  (sheet "Original" — NAME + DATE/AMOUNT PAID)

Override:
  python scripts/annotate_joined_date_review.py --file "data/your-file.xlsx"

Output: data/joined_dates - review.xlsx
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DEFAULT_INPUT = DATA / "Members List as of 11-01-2024 Including Original.xlsx"
OUTPUT_XLSX = DATA / "joined_dates - review.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))
from seed_from_exports import load_env_file, require_openpyxl, require_psycopg

load_env_file()

CONNECT_TIMEOUT_SEC = 15


def norm_name(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text.strip().lower())
    text = re.sub(r"\s*\(deceased\)\s*", " ", text, flags=re.I)
    text = re.sub(r"^dr\.?\s+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def primary_household(full_name: str | None) -> str:
    if not full_name:
        return ""
    return norm_name(str(full_name).split("/")[0].strip())


def collapse_name(value) -> str:
    return re.sub(r"[^a-z0-9]", "", norm_name(value))


def parse_excel_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Collapse typos like 8/25//2019
    text = re.sub(r"/+", "/", text)
    # Fix mangled Excel text like 4/12012 or 6/182017 (missing slash before year)
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})(\d{4})", text)
    if m:
        month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(year, month, day)
        except ValueError:
            return None
    # Truncated year text like 6/30/201 (meant 6/30/2019 — last digit dropped)
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{3})", text)
    if m:
        month, day, y3 = int(m.group(1)), int(m.group(2)), m.group(3)
        # Prefer completing with 9 first (201→2019 matches Hibret sheets), then 0-8
        for suffix in ("9", "0", "1", "2", "3", "4", "5", "6", "7", "8"):
            try:
                return date(int(y3 + suffix), month, day)
            except ValueError:
                continue
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%b %d %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def member_keys(row: dict) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    paypal = norm_name(row.get("paypal_name"))
    full_primary = primary_household(row.get("full_name"))
    first_last = norm_name(f"{row.get('first_name') or ''} {row.get('last_name') or ''}".strip())
    full = norm_name(row.get("full_name"))

    if paypal:
        keys.append((paypal, "paypal"))
        collapsed = collapse_name(row.get("paypal_name"))
        if collapsed:
            keys.append((f"__collapsed__:{collapsed}", "paypal_fuzzy"))
    if full_primary and full_primary != paypal:
        keys.append((full_primary, "full_name"))
        collapsed = collapse_name(full_primary)
        if collapsed:
            keys.append((f"__collapsed__:{collapsed}", "full_name_fuzzy"))
    if first_last and first_last not in {paypal, full_primary}:
        keys.append((first_last, "first_last"))
    if full and full not in {k[0] for k in keys if not k[0].startswith("__collapsed__")}:
        keys.append((full, "full_name"))
    # Also index spouse side of slash names
    if row.get("full_name") and "/" in str(row.get("full_name")):
        spouse = norm_name(str(row.get("full_name")).split("/", 1)[1].strip())
        if spouse:
            keys.append((spouse, "spouse_side"))
            keys.append((f"__collapsed__:{collapse_name(spouse)}", "spouse_fuzzy"))
    return keys


def build_member_index(rows: list[dict]) -> dict[str, list[dict]]:
    index: dict[str, list[dict]] = {}
    for row in rows:
        seen: set[str] = set()
        for key, _mt in member_keys(row):
            if not key or key in seen:
                continue
            seen.add(key)
            index.setdefault(key, []).append(row)
    return index


def pick_best_match(candidates: list[dict]) -> tuple[dict | None, str]:
    if not candidates:
        return None, "no_match"
    active = [c for c in candidates if str(c.get("status") or "").lower() == "active"]
    pool = active or candidates
    pool = sorted(pool, key=lambda c: (c.get("member_number") is None, c.get("member_number") or 9999))
    if len({c["id"] for c in pool}) > 1:
        return pool[0], "review"
    return pool[0], "exact"


def load_db_members() -> list[dict]:
    psycopg2 = require_psycopg()
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise SystemExit("DATABASE_URL missing in .env")
    kwargs = {"connect_timeout": CONNECT_TIMEOUT_SEC}
    if "render.com" in url:
        kwargs["sslmode"] = "require"
    conn = psycopg2.connect(url, **kwargs)
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, member_number, first_name, last_name, full_name, paypal_name,
                   spouse_name, email, status, joined_date
            FROM members
            ORDER BY member_number NULLS LAST, id
            """
        )
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.close()
        return rows
    finally:
        conn.close()


def cell_at(row, idx):
    if not row or idx is None or idx >= len(row):
        return None
    return row[idx]


def joined_from_row_or_next(rows: list, i: int, date_cols: list[int]):
    """
    Hibret member lists often put the date one blank row below the name
    (merged/tall cells). Prefer same-row date; else use next blank row's date.
    """
    r = rows[i]
    for col in date_cols:
        joined = parse_excel_date(cell_at(r, col))
        if joined:
            return joined, cell_at(r, col)
    if i + 1 < len(rows):
        nxt = rows[i + 1]
        # Only borrow from a blank "spacer" row (no member number / name)
        next_no = cell_at(nxt, 0)
        next_name = cell_at(nxt, 1) if len(nxt) > 1 else None
        if next_no is None and not next_name:
            for col in date_cols:
                joined = parse_excel_date(cell_at(nxt, col))
                if joined:
                    return joined, cell_at(nxt, col)
    return None, None


def load_excel_joined_rows(path: Path, openpyxl) -> list[dict]:
    if not path.exists():
        raise SystemExit(f"Missing Excel: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Prefer Original sheet (has registration dates); else first sheet with Date-ish headers
    sheet_name = None
    for candidate in ("Original", "Active List", "Sheet1"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out: list[dict] = []
    if sheet_name == "Original":
        for i, r in enumerate(rows):
            if not r:
                continue
            no, name = r[0], r[1] if len(r) > 1 else None
            if not isinstance(no, (int, float)) or not name:
                continue
            joined, raw = joined_from_row_or_next(rows, i, [6, 7, 2])
            out.append(
                {
                    "excel_no": int(no),
                    "excel_name": str(name).replace("\n", " ").strip(),
                    "excel_name_norm": primary_household(str(name)),
                    "excel_joined": joined,
                    "excel_joined_raw": raw,
                    "sheet": sheet_name,
                }
            )
        return out

    # Generic: detect header row with name + date columns
    header_idx = None
    name_col = date_col = no_col = None
    for i, r in enumerate(rows[:30]):
        if not r:
            continue
        labels = [str(c or "").strip().lower().replace("\n", " ") for c in r]
        for j, lab in enumerate(labels):
            if lab in ("name", "full name", "name with spouses", "recipient first name") and name_col is None:
                name_col = j
            if "join" in lab or lab in ("date", "date joined", "joined", "joined date", "registration date"):
                date_col = j
            if lab in ("no", "no.", "#", "member #", "member number"):
                no_col = j
            if "date/" in lab and "paid" in lab and date_col is None:
                date_col = j
        if name_col is not None and date_col is not None:
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit(
            f"Could not find Name + Joined-date columns in {path.name} sheet '{sheet_name}'. "
            "Put columns like Name / Joined date, or use the Original members list."
        )

    date_cols = [date_col]
    # Also try common date columns if header detection is narrow
    for extra in (2, 6, 7):
        if extra not in date_cols:
            date_cols.append(extra)

    for i in range(header_idx + 1, len(rows)):
        r = rows[i]
        if not r or name_col >= len(r):
            continue
        name = r[name_col]
        if not name:
            continue
        no = None
        if no_col is not None and no_col < len(r) and isinstance(r[no_col], (int, float)):
            no = int(r[no_col])
        joined, raw = joined_from_row_or_next(rows, i, date_cols)

        # Pattern: (None, Name, None) then (No, None, Date) — number/date shifted one row down
        if joined is None and no is None and i + 1 < len(rows):
            nxt = rows[i + 1]
            next_no = cell_at(nxt, no_col if no_col is not None else 0)
            next_name = cell_at(nxt, name_col)
            if isinstance(next_no, (int, float)) and not next_name:
                no = int(next_no)
                for col in date_cols:
                    joined = parse_excel_date(cell_at(nxt, col))
                    if joined:
                        raw = cell_at(nxt, col)
                        break

        out.append(
            {
                "excel_no": no,
                "excel_name": str(name).replace("\n", " ").strip(),
                "excel_name_norm": primary_household(str(name)),
                "excel_joined": joined,
                "excel_joined_raw": raw,
                "sheet": sheet_name,
            }
        )
    return out


def find_match(excel_row: dict, by_number: dict[int, dict], index: dict[str, list[dict]]) -> tuple[dict | None, str]:
    no = excel_row.get("excel_no")
    if no is not None and no in by_number:
        return by_number[no], "member_number"

    key = excel_row["excel_name_norm"]
    collapsed = f"__collapsed__:{collapse_name(excel_row.get('excel_name'))}"
    for lookup in (key, collapsed, norm_name(excel_row.get("excel_name"))):
        if lookup and lookup in index:
            return pick_best_match(index[lookup])
    # Try first segment only if slash name already used primary
    return None, "no_match"


def action_for(match: dict | None, excel_joined: date | None, match_type: str) -> str:
    if not excel_joined:
        return "skip_no_excel_date"
    if match_type == "no_match" or not match:
        return "skip_no_match"
    if match_type == "review":
        return "review"
    current = match.get("joined_date")
    if current is None:
        return "update"
    if isinstance(current, datetime):
        current = current.date()
    if current == excel_joined:
        return "already_same"
    return "review_overwrite"


def main() -> int:
    parser = argparse.ArgumentParser(description="Annotate joined-date Excel against CRM")
    parser.add_argument("--file", type=Path, default=DEFAULT_INPUT, help="Input Excel path")
    args = parser.parse_args()

    openpyxl = require_openpyxl()
    excel_rows = load_excel_joined_rows(args.file, openpyxl)
    db_members = load_db_members()
    by_number = {
        int(m["member_number"]): m
        for m in db_members
        if m.get("member_number") is not None
    }
    index = build_member_index(db_members)

    review_rows = []
    counts: dict[str, int] = {}
    for er in excel_rows:
        match, match_type = find_match(er, by_number, index)
        action = action_for(match, er.get("excel_joined"), match_type)
        counts[action] = counts.get(action, 0) + 1
        counts[f"match:{match_type}"] = counts.get(f"match:{match_type}", 0) + 1
        db_joined = match.get("joined_date") if match else None
        if isinstance(db_joined, datetime):
            db_joined = db_joined.date()
        review_rows.append(
            {
                "excel_no": er.get("excel_no"),
                "excel_name": er.get("excel_name"),
                "excel_joined": er.get("excel_joined").isoformat() if er.get("excel_joined") else "",
                "excel_joined_raw": str(er.get("excel_joined_raw") or ""),
                "db_member_number": match.get("member_number") if match else "",
                "db_paypal_name": match.get("paypal_name") if match else "",
                "db_full_name": match.get("full_name") if match else "",
                "db_status": match.get("status") if match else "",
                "db_joined_date": db_joined.isoformat() if db_joined else "",
                "match_type": match_type,
                "action": action,
                "your_ok": "Y" if action == "update" else ("N" if action.startswith("skip") else ""),
            }
        )

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    headers = [
        "excel_no",
        "excel_name",
        "excel_joined",
        "excel_joined_raw",
        "db_member_number",
        "db_paypal_name",
        "db_full_name",
        "db_status",
        "db_joined_date",
        "match_type",
        "action",
        "your_ok",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in review_rows:
        ws.append([row[h] for h in headers])

    note = wb.create_sheet("README", 0)
    note["A1"] = "Joined-date import review"
    note["A2"] = f"Source: {args.file}"
    note["A3"] = "Set your_ok=Y on rows to apply. action=update are pre-checked Y."
    note["A4"] = "action=review / review_overwrite need your confirmation (leave blank until checked)."
    note["A5"] = "Then: python scripts/apply_joined_date_updates.py          # dry-run"
    note["A6"] = "      python scripts/apply_joined_date_updates.py --apply"
    note["A8"] = "Counts:"
    r = 9
    for k, v in sorted(counts.items()):
        note[f"A{r}"] = f"{k}: {v}"
        r += 1

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")
    print("Counts:", dict(sorted(counts.items())))
    print(f"Excel rows: {len(excel_rows)} | CRM members: {len(db_members)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
