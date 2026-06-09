#!/usr/bin/env python3
"""
Import Hibret Edir member and invoice exports into PostgreSQL.

Expected local files (gitignored — never commit):
  data/Members List as of 11-01-2024 Including Original.xlsx
  data/PayPal_Invoice_Batch_Upload_Template V4.csv
  data/unpaid_invoices.xlsx  OR  data/unpaid_invoices.csv  OR  data/Download-*.csv (PayPal export)

Usage:
  Add DATABASE_URL to .env (or set in shell), then:
  python scripts/seed_from_exports.py --dry-run
  python scripts/seed_from_exports.py
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
ENV_PATH = ROOT / ".env"
PUBLIC_STATS_JSON = ROOT / "public" / "member-stats.json"


def load_env_file() -> None:
    """Load .env into os.environ (does not override existing vars)."""
    if not ENV_PATH.exists():
        return
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, _, value = stripped.partition("=")
        key = key.strip()
        value = value.strip()
        if (value.startswith('"') and value.endswith('"')) or (
            value.startswith("'") and value.endswith("'")
        ):
            value = value[1:-1]
        value = value.replace("postgressql://", "postgresql://", 1)
        if key and key not in os.environ:
            os.environ[key] = value


load_env_file()

MEMBERS_XLSX = DATA / "Members List as of 11-01-2024 Including Original.xlsx"
MEMBER_CAP = 200
PAYPAL_CSV = DATA / "PayPal_Invoice_Batch_Upload_Template V4.csv"

# Mistaken / non-invoice exports — never use for seeding
IGNORED_INVOICE_FILES = {
    "Unpaid_Invoice to follow up 2024-2025 .xlsx",
    "PayPal_Invoice_Batch_Upload_Template V4.csv",
}

INVOICE_CANDIDATES = (
    DATA / "unpaid_invoices.xlsx",
    DATA / "unpaid_invoices.csv",
    DATA / "unpaid_invoices.json",
    DATA / "invoices.xlsx",
    DATA / "invoices.csv",
    DATA / "invoices.json",
)

CONNECT_TIMEOUT_SEC = 10


def require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

    return openpyxl


def require_psycopg():
    try:
        import psycopg2  # noqa: F401
    except ImportError:
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "psycopg2-binary", "-q"])
    import psycopg2

    return psycopg2


def normalize_phone(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "none":
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) < 10:
        return text
    last10 = digits[-10:]
    return f"{last10[:3]}-{last10[3:6]}-{last10[6:]}"


def clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_event_number(item_name: str | None) -> int | None:
    if not item_name:
        return None
    match = re.search(r"#\s*(\d+)", str(item_name))
    return int(match.group(1)) if match else None


def parse_deceased_name(item_name: str | None) -> str | None:
    if not item_name:
        return None
    text = str(item_name).strip()
    text = re.sub(r"^#\s*\d+\s*", "", text).strip()
    return text or None


def paypal_id_from_link(link: str | None) -> str | None:
    if not link:
        return None
    text = str(link).strip()
    if "#" in text:
        return text.rsplit("#", 1)[-1].strip() or None
    return text or None


def load_original_details(openpyxl):
    wb = openpyxl.load_workbook(MEMBERS_XLSX, read_only=True, data_only=True)
    ws = wb["Original"]
    by_number: dict[int, dict] = {}
    by_name: dict[str, dict] = {}

    current_no = None
    current_name = None
    address_lines: list[str] = []

    def flush():
        nonlocal current_no, current_name, address_lines
        if current_no is None and not current_name:
            return
        email = None
        home = None
        mobile = None
        address = None

        rows = address_lines[:]
        address_lines = []

        if rows:
            first = rows[0]
            if len(first) > 5 and first[5] and "@" in str(first[5]):
                email = clean_text(first[5])
            if len(first) > 3:
                home = normalize_phone(first[3])
            if len(first) > 4:
                mobile = normalize_phone(first[4])
            if len(first) > 2 and first[2]:
                parts = [clean_text(first[2])]
                if len(rows) > 1 and rows[1][2]:
                    parts.append(clean_text(rows[1][2]))
                address = ", ".join(p for p in parts if p)

        payload = {
            "email": email,
            "home_phone": home,
            "mobile": mobile,
            "address": address,
        }
        if current_no is not None:
            by_number[int(current_no)] = payload
        if current_name:
            by_name[current_name.lower()] = payload

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        no_cell = row[0]
        name_cell = row[1]

        if isinstance(no_cell, (int, float)) and name_cell:
            flush()
            current_no = int(no_cell)
            current_name = clean_text(name_cell)
            address_lines = [row]
            continue

        if current_name and row[2] and not isinstance(no_cell, (int, float)):
            address_lines.append(row)

    flush()
    wb.close()
    return by_number, by_name


def load_paypal_csv_emails() -> dict[str, str]:
    emails: dict[str, str] = {}
    if not PAYPAL_CSV.exists():
        return emails

    with PAYPAL_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            email = clean_text(row.get("Recipient Email"))
            first = clean_text(row.get("Recipient First Name")) or ""
            last = clean_text(row.get("Recipient Last Name")) or ""
            if not email:
                continue
            full = f"{first} {last}".strip().lower()
            paypal_name = full
            if full:
                emails[full] = email
            if first and last:
                emails[f"{first.lower()}|{last.lower()}"] = email
    return emails


def lookup_email(paypal_name, first, last, by_number, by_name, csv_emails, member_number):
    details = by_number.get(member_number) if member_number is not None else None
    if not details and paypal_name:
        details = by_name.get(paypal_name.lower())
    if details and details.get("email"):
        return details["email"]

    full = f"{first or ''} {last or ''}".strip().lower()
    if full in csv_emails:
        return csv_emails[full]
    if first and last:
        key = f"{first.lower()}|{last.lower()}"
        if key in csv_emails:
            return csv_emails[key]
    if paypal_name and paypal_name.lower() in csv_emails:
        return csv_emails[paypal_name.lower()]
    return None


def parse_member_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = clean_text(value)
    if not text or not text.isdigit():
        return None
    return int(text)


def load_members(openpyxl):
    by_number_orig, by_name_orig = load_original_details(openpyxl)
    csv_emails = load_paypal_csv_emails()
    members: list[dict] = []

    wb = openpyxl.load_workbook(MEMBERS_XLSX, read_only=True, data_only=True)

    def add_member(payload: dict):
        members.append(payload)

    # Load inactive/deceased first so Active List wins on duplicate member_number upserts.
    ws = wb["Deceased or Out"]
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] not in ("Not Active", "Deceased"):
            continue
        status = "Deceased" if row[0] == "Deceased" else "Not Active"
        member_number = parse_member_number(row[2]) if len(row) > 2 else None
        first = clean_text(row[4]) if len(row) > 4 else None
        last = clean_text(row[5]) if len(row) > 5 else None
        full_name = clean_text(row[6]) if len(row) > 6 else None
        paypal_name = clean_text(row[7]) if len(row) > 7 else full_name
        home = normalize_phone(row[8]) if len(row) > 8 else None
        mobile = normalize_phone(row[9]) if len(row) > 9 else None

        orig = by_number_orig.get(member_number, {}) if member_number is not None else {}
        email = lookup_email(
            paypal_name, first, last, by_number_orig, by_name_orig, csv_emails, member_number
        )

        add_member(
            {
                "member_number": member_number,
                "status": status,
                "first_name": first,
                "last_name": last,
                "full_name": full_name,
                "paypal_name": paypal_name,
                "email": email,
                "mobile": mobile or orig.get("mobile"),
                "home_phone": home or orig.get("home_phone"),
                "address": orig.get("address"),
                "notes": "",
            }
        )

    ws = wb["Active List"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        status = clean_text(row[0]) or "Active"
        paypal_name = clean_text(row[1])
        member_number = parse_member_number(row[2])
        first = clean_text(row[3])
        last = clean_text(row[4])
        full_name = clean_text(row[5]) or paypal_name
        home = normalize_phone(row[6])
        mobile = normalize_phone(row[7])

        orig = by_number_orig.get(member_number, {}) if member_number is not None else {}
        if not orig and full_name:
            orig = by_name_orig.get(full_name.lower(), {})

        email = lookup_email(
            paypal_name, first, last, by_number_orig, by_name_orig, csv_emails, member_number
        )

        add_member(
            {
                "member_number": member_number,
                "status": status,
                "first_name": first,
                "last_name": last,
                "full_name": full_name,
                "paypal_name": paypal_name,
                "email": email,
                "mobile": mobile or orig.get("mobile"),
                "home_phone": home or orig.get("home_phone"),
                "address": orig.get("address"),
                "notes": "",
            }
        )

    wb.close()
    return members


def parse_us_date(value) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def normalize_paypal_status(value: str | None) -> str:
    text = (value or "").lower()
    if "paid" in text and "unpaid" not in text:
        return "Paid"
    if "unpaid" in text or "sent" in text:
        return "Unpaid"
    if "cancel" in text:
        return "Cancelled"
    return clean_text(value) or "Unpaid"


def find_invoice_export() -> Path | None:
    for path in INVOICE_CANDIDATES:
        if path.exists() and path.name not in IGNORED_INVOICE_FILES:
            return path

    # PayPal invoice exports (e.g. Download-1780537524021.csv)
    downloads = sorted(DATA.glob("Download*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in downloads:
        if path.name not in IGNORED_INVOICE_FILES:
            return path

    return None


def is_paypal_invoice_export(rows: list[dict]) -> bool:
    if not rows:
        return False
    sample = rows[0]
    return "PayPal Invoice ID" in sample and "Invoice number" in sample and "Type" in sample


def load_paypal_export_csv(path: Path) -> list[dict]:
    invoices: list[dict] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if clean_text(row.get("Type")) != "Invoice item":
                continue

            invoice_number = row.get("Invoice number")
            name = clean_text(row.get("Name"))
            item_name = clean_text(row.get("Item name"))
            if not invoice_number or not name or not item_name:
                continue

            paypal_id = clean_text(row.get("PayPal Invoice ID"))
            paypal_link = f"https://www.paypal.com/invoice/p/#{paypal_id}" if paypal_id else None
            amount = float(row.get("Item total") or row.get("Item unit price") or 110)
            amount_due_raw = row.get("Amount due")
            amount_due = float(amount_due_raw) if amount_due_raw not in (None, "") else amount
            status = normalize_paypal_status(row.get("Status"))

            invoices.append(
                {
                    "invoice_number": int(float(invoice_number)),
                    "paypal_invoice_id": paypal_id,
                    "paypal_name": name,
                    "email": clean_text(row.get("To email")),
                    "event_number": parse_event_number(item_name),
                    "deceased_name": parse_deceased_name(item_name),
                    "item_name": item_name,
                    "status": status,
                    "amount": amount,
                    "amount_due": amount_due if status != "Paid" else 0,
                    "sent_date": parse_us_date(row.get("Date created") or row.get("Invoice date")),
                    "paid_date": parse_us_date(row.get("Last payment date")) if status == "Paid" else None,
                    "payment_method": "PayPal",
                    "paypal_link": paypal_link,
                    "notes": clean_text(row.get("Customer note")),
                }
            )

    return invoices


def normalize_invoice_row(row: dict) -> dict | None:
    invoice_number = (
        row.get("invoice_number")
        or row.get("invoice_num")
        or row.get("Invoice number")
        or row.get("Invoice Number")
    )
    name = (
        row.get("paypal_name")
        or row.get("name")
        or row.get("Name")
        or row.get("recipient")
    )
    if not invoice_number or not name:
        return None

    item_name = row.get("item_name") or row.get("item") or row.get("Item name") or row.get("Item Name")
    paypal_link = (
        row.get("paypal_link")
        or row.get("link")
        or row.get("Invoice shipping address")
        or row.get("invoice_link")
    )
    email = row.get("email") or row.get("To email") or row.get("member_email")
    sent_date = row.get("sent_date") or row.get("date") or row.get("Date created")
    if isinstance(sent_date, datetime):
        sent_date = sent_date.date().isoformat()
    elif isinstance(sent_date, date):
        sent_date = sent_date.isoformat()
    else:
        sent_date = clean_text(sent_date)

    status = clean_text(row.get("status") or row.get("Status")) or "Unpaid"
    amount = row.get("amount") or row.get("total") or row.get("Amount") or 110
    amount_due = row.get("amount_due") if row.get("amount_due") is not None else amount

    return {
        "invoice_number": int(float(invoice_number)),
        "paypal_invoice_id": paypal_id_from_link(clean_text(paypal_link)),
        "paypal_name": clean_text(name),
        "email": clean_text(email),
        "event_number": parse_event_number(clean_text(item_name)),
        "deceased_name": parse_deceased_name(clean_text(item_name)),
        "item_name": clean_text(item_name),
        "status": status,
        "amount": float(amount),
        "amount_due": float(amount_due),
        "sent_date": sent_date,
        "paid_date": clean_text(row.get("paid_date") or row.get("Paid date")),
        "payment_method": clean_text(row.get("payment_method") or row.get("method")) or "PayPal",
        "paypal_link": clean_text(paypal_link),
        "notes": clean_text(row.get("notes") or row.get("Note")),
    }


def load_invoices_from_csv(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))

    if is_paypal_invoice_export(rows):
        return load_paypal_export_csv(path)

    invoices: list[dict] = []
    for row in rows:
        normalized = normalize_invoice_row(row)
        if normalized:
            invoices.append(normalized)
    return invoices


def load_invoices_from_json(path: Path) -> list[dict]:
    import json

    payload = json.loads(path.read_text(encoding="utf-8"))
    rows = payload if isinstance(payload, list) else payload.get("invoices", [])
    invoices: list[dict] = []
    for row in rows:
        normalized = normalize_invoice_row(row)
        if normalized:
            invoices.append(normalized)
    return invoices


def load_invoices_from_xlsx(openpyxl, path: Path) -> list[dict]:
    invoices: list[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = None
        headers: list[str] = []
        for idx, row in enumerate(rows[:5]):
            if not row:
                continue
            labels = [clean_text(cell) or "" for cell in row]
            if any(label.lower() in ("name", "invoice number", "item name") for label in labels):
                header_idx = idx
                headers = labels
                break

        if header_idx is None:
            continue

        for row in rows[header_idx + 1 :]:
            if not row:
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_dict[header] = row[i]
            normalized = normalize_invoice_row(row_dict)
            if normalized:
                invoices.append(normalized)

    wb.close()
    return invoices


def load_unpaid_invoices(openpyxl) -> list[dict]:
    path = find_invoice_export()
    if not path:
        print("No invoice export found — seeding members only.")
        print("When ready, add: data/unpaid_invoices.xlsx or data/unpaid_invoices.csv")
        return []

    print(f"Using invoice export: {path.name}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_invoices_from_csv(path)
    if suffix == ".json":
        return load_invoices_from_json(path)
    if suffix in (".xlsx", ".xls"):
        return load_invoices_from_xlsx(openpyxl, path)
    print(f"Unsupported invoice file type: {path}")
    return []


def connect(db_url: str):
    psycopg2 = require_psycopg()
    kwargs: dict = {"connect_timeout": CONNECT_TIMEOUT_SEC}
    if "render.com" in db_url:
        kwargs["sslmode"] = "require"
    return psycopg2.connect(db_url, **kwargs)


def upsert_member_without_number(cur, member: dict) -> bool:
    email = member.get("email")
    full_name = member.get("full_name")
    if email:
        cur.execute(
            "SELECT id FROM members WHERE LOWER(email) = LOWER(%s) LIMIT 1",
            (email,),
        )
    elif full_name:
        cur.execute(
            """
            SELECT id FROM members
            WHERE member_number IS NULL AND LOWER(full_name) = LOWER(%s)
            LIMIT 1
            """,
            (full_name,),
        )
    else:
        return False

    row = cur.fetchone()
    fields = (
        "status", "first_name", "last_name", "full_name", "paypal_name",
        "email", "mobile", "home_phone", "address", "notes",
    )
    if row:
        cur.execute(
            f"""
            UPDATE members SET
              status = %(status)s,
              first_name = %(first_name)s,
              last_name = %(last_name)s,
              full_name = %(full_name)s,
              paypal_name = %(paypal_name)s,
              email = %(email)s,
              mobile = %(mobile)s,
              home_phone = %(home_phone)s,
              address = %(address)s,
              notes = %(notes)s,
              updated_at = NOW()
            WHERE id = %(id)s
            """,
            {**member, "id": row[0]},
        )
        return True

    cur.execute(
        """
        INSERT INTO members (
          member_number, status, first_name, last_name, full_name, paypal_name,
          email, mobile, home_phone, address, notes
        ) VALUES (
          NULL, %(status)s, %(first_name)s, %(last_name)s, %(full_name)s, %(paypal_name)s,
          %(email)s, %(mobile)s, %(home_phone)s, %(address)s, %(notes)s
        )
        """,
        member,
    )
    return True


def upsert_members(conn, members: list[dict], dry_run: bool) -> int:
    sql = """
        INSERT INTO members (
          member_number, status, first_name, last_name, full_name, paypal_name,
          email, mobile, home_phone, address, notes
        ) VALUES (
          %(member_number)s, %(status)s, %(first_name)s, %(last_name)s, %(full_name)s, %(paypal_name)s,
          %(email)s, %(mobile)s, %(home_phone)s, %(address)s, %(notes)s
        )
        ON CONFLICT (member_number) DO UPDATE SET
          status = EXCLUDED.status,
          first_name = EXCLUDED.first_name,
          last_name = EXCLUDED.last_name,
          full_name = EXCLUDED.full_name,
          paypal_name = EXCLUDED.paypal_name,
          email = EXCLUDED.email,
          mobile = EXCLUDED.mobile,
          home_phone = EXCLUDED.home_phone,
          address = EXCLUDED.address,
          notes = EXCLUDED.notes,
          updated_at = NOW()
    """
    count = 0
    if dry_run:
        for member in members:
            if member.get("member_number") is None:
                print(f"[dry-run] member (no #): {member.get('paypal_name')}")
                count += 1
                continue
            print(f"[dry-run] member #{member['member_number']}: {member.get('paypal_name')}")
            count += 1
        return count

    with conn.cursor() as cur:
        for member in members:
            if member.get("member_number") is None:
                if upsert_member_without_number(cur, member):
                    count += 1
                continue
            cur.execute(sql, member)
            count += 1
        conn.commit()
    return count


def find_member_id(cur, invoice_name: str | None, email: str | None):
    """Match invoice recipient to CRM member. PayPal name is primary; household full_name may list both spouses."""
    if email:
        cur.execute(
            "SELECT id FROM members WHERE LOWER(email) = LOWER(%s) LIMIT 1",
            (email,),
        )
        row = cur.fetchone()
        if row:
            return row[0]

    if not invoice_name:
        return None

    name = invoice_name.strip()

    cur.execute(
        "SELECT id FROM members WHERE LOWER(TRIM(paypal_name)) = LOWER(%s) LIMIT 1",
        (name,),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    cur.execute(
        "SELECT id FROM members WHERE LOWER(TRIM(full_name)) = LOWER(%s) LIMIT 1",
        (name,),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    # Household names: "Primary/Spouse" — match either side or surviving spouse on same member #.
    cur.execute(
        """
        SELECT id FROM members
        WHERE LOWER(TRIM(split_part(full_name, '/', 1))) = LOWER(%s)
           OR LOWER(TRIM(split_part(full_name, '/', 2))) = LOWER(%s)
           OR LOWER(full_name) LIKE LOWER(%s) || '/%'
           OR LOWER(full_name) LIKE '%/' || LOWER(%s)
        LIMIT 1
        """,
        (name, name, name, name),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    # Fuzzy: PayPal export name contained in paypal_name (deceased notes, c/o, etc.)
    cur.execute(
        """
        SELECT id FROM members
        WHERE LOWER(paypal_name) LIKE '%%' || LOWER(%s) || '%%'
           OR LOWER(full_name) LIKE '%%' || LOWER(%s) || '%%'
        ORDER BY
          CASE WHEN LOWER(TRIM(paypal_name)) = LOWER(%s) THEN 0
               WHEN LOWER(paypal_name) LIKE LOWER(%s) || '%%' THEN 1
               ELSE 2 END,
          id
        LIMIT 1
        """,
        (name, name, name, name),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    return None


def upsert_event(cur, event_number: int | None, deceased_name: str | None, dry_run: bool):
    if not event_number:
        return None
    if dry_run:
        return event_number
    cur.execute(
        """
        INSERT INTO events (event_number, deceased_name, status)
        VALUES (%s, %s, 'Active')
        ON CONFLICT (event_number) DO UPDATE SET
          deceased_name = COALESCE(EXCLUDED.deceased_name, events.deceased_name),
          updated_at = NOW()
        RETURNING id
        """,
        (event_number, deceased_name or f"Event #{event_number}"),
    )
    return cur.fetchone()[0]


def upsert_invoices(conn, invoices: list[dict], dry_run: bool):
    unmatched = 0
    count = 0
    if dry_run:
        for inv in invoices:
            print(
                f"[dry-run] invoice {inv.get('invoice_number')} event #{inv.get('event_number')} "
                f"{inv.get('paypal_name')}"
            )
            count += 1
        return count, unmatched

    with conn.cursor() as cur:
        for inv in invoices:
            member_id = find_member_id(cur, inv.get("paypal_name"), inv.get("email"))
            if not member_id:
                unmatched += 1
                print(
                    f"WARN no member match: invoice {inv.get('invoice_number')} "
                    f"{inv.get('paypal_name')} <{inv.get('email')}>"
                )

            event_id = upsert_event(cur, inv.get("event_number"), inv.get("deceased_name"), dry_run=False)

            if inv.get("paypal_invoice_id"):
                cur.execute(
                    """
                    INSERT INTO invoices (
                      paypal_invoice_id, invoice_number, member_id, event_id, status,
                      amount, amount_due, sent_date, paid_date, payment_method, paypal_link,
                      recipient_name
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (paypal_invoice_id) DO UPDATE SET
                      invoice_number = EXCLUDED.invoice_number,
                      member_id = EXCLUDED.member_id,
                      event_id = EXCLUDED.event_id,
                      status = EXCLUDED.status,
                      amount = EXCLUDED.amount,
                      amount_due = EXCLUDED.amount_due,
                      sent_date = EXCLUDED.sent_date,
                      paid_date = EXCLUDED.paid_date,
                      payment_method = EXCLUDED.payment_method,
                      paypal_link = EXCLUDED.paypal_link,
                      recipient_name = EXCLUDED.recipient_name,
                      updated_at = NOW()
                    """,
                    (
                        inv.get("paypal_invoice_id"),
                        inv.get("invoice_number"),
                        member_id,
                        event_id,
                        inv.get("status"),
                        inv.get("amount"),
                        inv.get("amount_due"),
                        inv.get("sent_date"),
                        inv.get("paid_date"),
                        inv.get("payment_method"),
                        inv.get("paypal_link"),
                        inv.get("paypal_name"),
                    ),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO invoices (
                      invoice_number, member_id, event_id, status,
                      amount, amount_due, sent_date, paid_date, payment_method, paypal_link,
                      recipient_name
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (invoice_number) WHERE invoice_number IS NOT NULL DO UPDATE SET
                      member_id = EXCLUDED.member_id,
                      event_id = EXCLUDED.event_id,
                      status = EXCLUDED.status,
                      amount = EXCLUDED.amount,
                      amount_due = EXCLUDED.amount_due,
                      sent_date = EXCLUDED.sent_date,
                      paid_date = EXCLUDED.paid_date,
                      payment_method = EXCLUDED.payment_method,
                      paypal_link = EXCLUDED.paypal_link,
                      recipient_name = EXCLUDED.recipient_name,
                      updated_at = NOW()
                    """,
                    (
                        inv.get("invoice_number"),
                        member_id,
                        event_id,
                        inv.get("status"),
                        inv.get("amount"),
                        inv.get("amount_due"),
                        inv.get("sent_date"),
                        inv.get("paid_date"),
                        inv.get("payment_method"),
                        inv.get("paypal_link"),
                        inv.get("paypal_name"),
                    ),
                )
            count += 1
        conn.commit()
    return count, unmatched


def ensure_files():
    missing = [p for p in (MEMBERS_XLSX, PAYPAL_CSV) if not p.exists()]
    if missing:
        print("Missing required member export files:")
        for path in missing:
            print(f"  - {path}")
        sys.exit(1)


def write_public_member_stats(members: list[dict]) -> dict:
    active = sum(1 for m in members if m.get("status") == "Active")
    inactive = len(members) - active
    payload = {
        "updated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source": MEMBERS_XLSX.name,
        "active_count": active,
        "total_count": len(members),
        "inactive_count": inactive,
        "member_cap": MEMBER_CAP,
    }
    PUBLIC_STATS_JSON.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return payload


def main():
    parser = argparse.ArgumentParser(description="Seed Hibret Edir DB from local exports")
    parser.add_argument("--dry-run", action="store_true", help="Parse files only, no DB writes")
    parser.add_argument(
        "--stats-only",
        action="store_true",
        help="Write public/member-stats.json from membership export (no DB)",
    )
    args = parser.parse_args()

    ensure_files()
    openpyxl = require_openpyxl()

    members = load_members(openpyxl)
    stats = write_public_member_stats(members)
    print(f"Public stats written: {PUBLIC_STATS_JSON.name} ({stats['active_count']} active)")

    if args.stats_only:
        print(
            f"Members: {stats['total_count']} total "
            f"({stats['active_count']} active, {stats['inactive_count']} inactive/out)"
        )
        return

    invoices = load_unpaid_invoices(openpyxl)

    active = stats["active_count"]
    with_email = sum(1 for m in members if m.get("email"))
    print(f"Members loaded: {len(members)} ({active} active, {with_email} with email)")
    if invoices:
        print(f"Unpaid invoices loaded: {len(invoices)}")
    else:
        print("Unpaid invoices: skipped")

    if args.dry_run:
        upsert_members(None, members, dry_run=True)
        if invoices:
            upsert_invoices(None, invoices, dry_run=True)
        print("Dry run complete.")
        return

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("DATABASE_URL is not set. Add it to .env or your shell environment.")
        sys.exit(1)

    conn = connect(db_url)
    try:
        member_count = upsert_members(conn, members, dry_run=False)
        print(f"Members upserted: {member_count}")
        if invoices:
            invoice_count, unmatched = upsert_invoices(conn, invoices, dry_run=False)
            print(f"Invoices upserted: {invoice_count}")
            if unmatched:
                print(f"Invoices without member match: {unmatched}")
        else:
            print("Invoices upserted: 0 (no export file)")
        print("Seed complete.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
